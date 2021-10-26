from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
wb = load_workbook('Athletes.xlsx')
ws = wb.active

row_count = ws.max_row + 1
column_count = ws.max_column + 1

athNo = 11072
for row in range(athNo, athNo+1):
    for col in range(1, 4):
        char = get_column_letter(col)
        print(ws[char + str(row)].value)